import xlrd

from openpyxl import load_workbook

from django.utils import timezone

from .models import ExcelFile


def _get_result(before_set, after_set):
    difference = before_set - after_set
    if difference:
        return f"removed: {int(difference.pop())}"

    difference = after_set - before_set
    if difference:
        return f"added: {int(difference.pop())}"


def _get_columns_before_and_after_xlsx(sheet):
    required = {'before', 'after'}
    need_len = 2
    result = {}

    first_row = sheet[1]

    if len(first_row) <= 1:
        return {}

    for i, cell in enumerate(first_row, start=1):
        if not cell.value:
            break

        if not required:
            break

        if cell.value in required:
            result.update({cell.value: i})
            required.remove(cell.value)

    if len(result.keys()) != need_len:
        return {}

    result.update({'sheet': sheet})

    return result


def _get_processing_result_xlsx(required_columns):
    row_min = 2
    sheet = required_columns['sheet']

    before_column = required_columns['before']
    after_column = required_columns['after']

    before_tuple = next(
        sheet.iter_cols(min_col=before_column, max_col=before_column, min_row=row_min, values_only=True)
    )
    after_tuple = next(
        sheet.iter_cols(min_col=after_column, max_col=after_column, min_row=row_min, values_only=True)
    )

    before_set = {item for item in before_tuple if item}
    after_set = {item for item in after_tuple if item}

    return _get_result(before_set, after_set)


def _processing_xlsx_file(file_path):
    work_book = load_workbook(file_path)

    required_columns = None
    for sheet in work_book.worksheets:
        tmp = _get_columns_before_and_after_xlsx(sheet)
        if tmp:
            required_columns = tmp
            break
    return _get_processing_result_xlsx(required_columns)


def _get_columns_before_and_after_xls(sheet):
    required = {'before', 'after'}
    need_len = 2
    result = {}

    try:
        first_row = sheet.row(0)
    except IndexError:
        return {}

    if len(first_row) <= 1:
        return {}

    for i, cell in enumerate(first_row):
        if not cell.value:
            break

        if not required:
            break

        if cell.value in required:
            result.update({cell.value: i})
            required.remove(cell.value)

    if len(result.keys()) != need_len:
        return {}

    result.update({'sheet': sheet})

    return result


def _get_processing_result_xls(required_columns):
    sheet = required_columns['sheet']
    before_column = required_columns['before']
    after_column = required_columns['after']

    before_set = {item.value for item in sheet.col(before_column) if item.value}
    after_set = {item.value for item in sheet.col(after_column) if item.value}

    before_set.discard('before')
    after_set.discard('after')

    return _get_result(before_set, after_set)


def _processing_xls_file(file_path):
    work_book = xlrd.open_workbook(file_path)

    required_columns = None
    for sheet in work_book.sheets():
        tmp = _get_columns_before_and_after_xls(sheet)
        if tmp:
            required_columns = tmp
            break
    return _get_processing_result_xls(required_columns)


def run_processing_excel_file(excel_file):
    if excel_file.processing_status == ExcelFile.ProcessingStatus.PROCESSED:
        return

    excel_file.processing_status = ExcelFile.ProcessingStatus.PROCESSING
    excel_file.save(update_fields=['processing_status', ])

    if excel_file.path.name.endswith('.xlsx'):
        result = _processing_xlsx_file(excel_file.path.path)
    else:  # mean '.xls'
        result = _processing_xls_file(excel_file.path.path)

    excel_file.processing_result = result
    excel_file.processing_status = ExcelFile.ProcessingStatus.PROCESSED
    excel_file.processing_stop = timezone.now()
    excel_file.save(
        update_fields=[
            'processing_result', 'processing_status', 'processing_stop'
        ]
    )
